import ipaddress
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Font
import re

count = 0
path = ''

# Custom subnet address management for 10.22.128.xx/18 SJC23
subnets_10_22_xx_xx = {
    ipaddress.ip_network('10.22.44.0/25'):[],
    ipaddress.ip_network('10.22.44.128/25'):[],
    ipaddress.ip_network('10.22.45.0/25'):[],
    ipaddress.ip_network('10.22.45.128/25'):[],
    ipaddress.ip_network('10.22.46.0/25'):[],
    ipaddress.ip_network('10.22.46.128/25'):[],
    ipaddress.ip_network('10.22.47.0/25'):[],
    ipaddress.ip_network('10.22.47.128/25'):[],
    ipaddress.ip_network('10.22.48.0/25'):[],
    ipaddress.ip_network('10.22.48.128/25'):[],
    ipaddress.ip_network('10.22.49.0/25'):[],
    ipaddress.ip_network('10.22.49.128/25'):[],
    ipaddress.ip_network('10.22.50.0/25'):[],
    ipaddress.ip_network('10.22.50.128/25'):[],
    ipaddress.ip_network('10.22.51.0/25'):[],
    ipaddress.ip_network('10.22.51.128/25'):[],
    ipaddress.ip_network('10.22.168.0/25'):[],
    ipaddress.ip_network('10.22.168.128/25'):[],
    ipaddress.ip_network('10.22.169.0/25'):[],
    ipaddress.ip_network('10.22.169.128/25'):[],
    ipaddress.ip_network('10.22.170.0/25'):[],
    ipaddress.ip_network('10.22.170.128/25'):[],
    ipaddress.ip_network('10.22.171.0/25'):[],
    ipaddress.ip_network('10.22.171.128/25'):[],
    ipaddress.ip_network('10.22.172.0/25'):[],
    ipaddress.ip_network('10.22.172.128/25'):[],
    ipaddress.ip_network('10.22.173.0/25'):[],
    ipaddress.ip_network('10.22.173.128/25'):[],
    ipaddress.ip_network('10.22.174.0/25'):[],
    ipaddress.ip_network('10.22.174.128/25'):[],
    ipaddress.ip_network('10.22.175.0/25'):[],
    ipaddress.ip_network('10.22.175.128/25'):[]
}

# Custom subnet address management for 10.29.xx.xx/18 SJCJ
subnets_10_29_xx_xx = { 
    ipaddress.ip_network('10.29.192.0/22'):[],
    ipaddress.ip_network('10.29.196.0/24'):[],
    ipaddress.ip_network('10.29.197.0/24'):[],
    ipaddress.ip_network('10.29.198.0/24'):[],
    ipaddress.ip_network('10.29.199.0/24'):[],
    ipaddress.ip_network('10.29.200.0/23'):[],
    ipaddress.ip_network('10.29.202.0/25'):[],
    ipaddress.ip_network('10.29.202.128/25'):[],
    ipaddress.ip_network('10.29.203.0/24'):[],
    ipaddress.ip_network('10.29.204.0/23'):[],
    ipaddress.ip_network('10.29.206.0/23'):[],
    ipaddress.ip_network('10.29.208.0/22'):[],
    ipaddress.ip_network('10.29.212.0/26'):[],
    ipaddress.ip_network('10.29.212.64/26'):[],
    ipaddress.ip_network('10.29.212.128/26'):[],
    ipaddress.ip_network('10.29.212.192/26'):[],
    ipaddress.ip_network('10.29.213.0/26'):[],
    ipaddress.ip_network('10.29.213.64/26'):[],
    ipaddress.ip_network('10.29.213.128/28'):[],
    ipaddress.ip_network('10.29.213.144/28'):[],
    ipaddress.ip_network('10.29.213.160/28'):[],
    ipaddress.ip_network('10.29.213.176/30'):[],
    ipaddress.ip_network('10.29.213.180/30'):[],
    ipaddress.ip_network('10.29.213.184/29'):[],
    ipaddress.ip_network('10.29.213.192/26'):[],
    ipaddress.ip_network('10.29.214.0/26'):[],
    ipaddress.ip_network('10.29.214.64/26'):[],
    ipaddress.ip_network('10.29.214.128/25'):[],
    ipaddress.ip_network('10.29.215.0/25'):[],
    ipaddress.ip_network('10.29.215.128/25'):[],
    ipaddress.ip_network('10.29.216.0/24'):[],
    ipaddress.ip_network('10.29.217.0/25'):[],
    ipaddress.ip_network('10.29.217.128/25'):[],
    ipaddress.ip_network('10.29.218.0/24'):[],
    ipaddress.ip_network('10.29.219.0/24'):[],
    ipaddress.ip_network('10.29.220.0/24'):[],
    ipaddress.ip_network('10.29.221.0/26'):[],
    ipaddress.ip_network('10.29.221.64/26'):[],
    ipaddress.ip_network('10.29.221.128/25'):[],
    ipaddress.ip_network('10.29.222.0/23'):[],
    ipaddress.ip_network('10.29.224.0/20'):[],
    ipaddress.ip_network('10.29.240.0/21'):[],
    ipaddress.ip_network('10.29.248.0/23'):[],
    ipaddress.ip_network('10.29.250.0/24'):[],
    ipaddress.ip_network('10.29.251.0/24'):[],
    ipaddress.ip_network('10.29.252.0/24'):[],
    ipaddress.ip_network('10.29.253.0/24'):[],
    ipaddress.ip_network('10.29.254.0/24'):[],
    ipaddress.ip_network('10.29.255.0/25'):[],
    ipaddress.ip_network('10.29.255.128/27'):[],
    ipaddress.ip_network('10.29.255.160/27'):[],
    ipaddress.ip_network('10.29.255.192/28'):[],
    ipaddress.ip_network('10.29.255.208/28'):[],
    ipaddress.ip_network('10.29.255.224/28'):[],
    ipaddress.ip_network('10.29.255.240/28'):[]
}

# Custom subnet address management for 171.68.xx.xx/19 SJCJ
subnets_171_68_xx_xx = {
    ipaddress.ip_network('171.68.165.128/26'):[],
    ipaddress.ip_network('171.68.165.192/29'):[],
    ipaddress.ip_network('171.68.165.200/29'):[],
    ipaddress.ip_network('171.68.165.224/27'):[],
    ipaddress.ip_network('171.68.173.0/26'):[],
    ipaddress.ip_network('171.68.173.64/26'):[],
    ipaddress.ip_network('171.68.173.128/25'):[]
}

# Custom subnet address management for 172.21.0.xx/16 SJCJ
subnets_172_21_xx_xx = {
    ipaddress.ip_network('172.21.48.0/25'):[],
    ipaddress.ip_network('172.21.48.128/28'):[],
    ipaddress.ip_network('172.21.48.144/28'):[],
    ipaddress.ip_network('172.21.48.160/28'):[],
    ipaddress.ip_network('172.21.48.176/28'):[],
    ipaddress.ip_network('172.21.48.192/26'):[],
    ipaddress.ip_network('172.21.49.0/27'):[],
    ipaddress.ip_network('172.21.49.32/27'):[],
    ipaddress.ip_network('172.21.49.64/27'):[],
    ipaddress.ip_network('172.21.49.96/27'):[],
    ipaddress.ip_network('172.21.49.128/28'):[],
    ipaddress.ip_network('172.21.49.144/28'):[],
    ipaddress.ip_network('172.21.49.160/27'):[],
    ipaddress.ip_network('172.21.49.192/27'):[],
    ipaddress.ip_network('172.21.49.224/27'):[],
    ipaddress.ip_network('172.21.50.0/27'):[],
    ipaddress.ip_network('172.21.50.32/27'):[],
    ipaddress.ip_network('172.21.50.64/28'):[],
    ipaddress.ip_network('172.21.50.80/28'):[],
    ipaddress.ip_network('172.21.50.96/27'):[],
    ipaddress.ip_network('172.21.50.128/27'):[],
    ipaddress.ip_network('172.21.50.160/27'):[],
    ipaddress.ip_network('172.21.50.192/27'):[],
    ipaddress.ip_network('172.21.50.224/27'):[],
    ipaddress.ip_network('172.21.51.0/24'):[],
    ipaddress.ip_network('172.21.52.0/29'):[],
    ipaddress.ip_network('172.21.52.8/29'):[],
    ipaddress.ip_network('172.21.52.16/28'):[],
    ipaddress.ip_network('172.21.52.32/27'):[],
    ipaddress.ip_network('172.21.52.64/26'):[],
    ipaddress.ip_network('172.21.52.128/26'):[],
    ipaddress.ip_network('172.21.52.192/27'):[],
    ipaddress.ip_network('172.21.52.224/27'):[],
    ipaddress.ip_network('172.21.53.0/28'):[],
    ipaddress.ip_network('172.21.53.16/28'):[],
    ipaddress.ip_network('172.21.53.32/30'):[],
    ipaddress.ip_network('172.21.53.36/30'):[],
    ipaddress.ip_network('172.21.53.40/29'):[],
    ipaddress.ip_network('172.21.53.48/28'):[],
    ipaddress.ip_network('172.21.53.64/26'):[],
    ipaddress.ip_network('172.21.53.128/27'):[],
    ipaddress.ip_network('172.21.53.160/27'):[],
    ipaddress.ip_network('172.21.53.192/27'):[],
    ipaddress.ip_network('172.21.53.224/27'):[],
    ipaddress.ip_network('172.21.54.0/25'):[],
    ipaddress.ip_network('172.21.54.128/27'):[],
    ipaddress.ip_network('172.21.54.160/27'):[],
    ipaddress.ip_network('172.21.54.192/29'):[],
    ipaddress.ip_network('172.21.54.200/29'):[],
    ipaddress.ip_network('172.21.54.208/28'):[],
    ipaddress.ip_network('172.21.54.224/27'):[],
    ipaddress.ip_network('172.21.55.0/24'):[],
    ipaddress.ip_network('172.21.56.0/26'):[],
    ipaddress.ip_network('172.21.56.64/26'):[],
    ipaddress.ip_network('172.21.56.128/26'):[],
    ipaddress.ip_network('172.21.56.192/27'):[],
    ipaddress.ip_network('172.21.56.224/27'):[],
    ipaddress.ip_network('172.21.57.0/26'):[],
    ipaddress.ip_network('172.21.57.64/27'):[],
    ipaddress.ip_network('172.21.57.96/27'):[],
    ipaddress.ip_network('172.21.57.128/27'):[],
    ipaddress.ip_network('172.21.57.160/27'):[],
    ipaddress.ip_network('172.21.57.192/27'):[],
    ipaddress.ip_network('172.21.57.224/27'):[],
    ipaddress.ip_network('172.21.58.0/25'):[],
    ipaddress.ip_network('172.21.58.128/26'):[],
    ipaddress.ip_network('172.21.58.192/28'):[],
    ipaddress.ip_network('172.21.58.208/28'):[],
    ipaddress.ip_network('172.21.58.224/27'):[],
    ipaddress.ip_network('172.21.59.0/26'):[],
    ipaddress.ip_network('172.21.59.64/27'):[],
    ipaddress.ip_network('172.21.59.96/27'):[],
    ipaddress.ip_network('172.21.59.128/27'):[],
    ipaddress.ip_network('172.21.59.160/27'):[],
    ipaddress.ip_network('172.21.59.192/27'):[],
    ipaddress.ip_network('172.21.59.224/27'):[],
    ipaddress.ip_network('172.21.60.0/26'):[],
    ipaddress.ip_network('172.21.60.64/26'):[],
    ipaddress.ip_network('172.21.60.128/25'):[],
    ipaddress.ip_network('172.21.61.0/27'):[],
    ipaddress.ip_network('172.21.61.32/27'):[],
    ipaddress.ip_network('172.21.61.64/27'):[],
    ipaddress.ip_network('172.21.61.96/27'):[],
    ipaddress.ip_network('172.21.61.128/27'):[],
    ipaddress.ip_network('172.21.61.160/27'):[],
    ipaddress.ip_network('172.21.61.192/26'):[],
    ipaddress.ip_network('172.21.62.0/25'):[],
    ipaddress.ip_network('172.21.62.128/25'):[],
    ipaddress.ip_network('172.21.63.0/26'):[],
    ipaddress.ip_network('172.21.63.64/27'):[],
    ipaddress.ip_network('172.21.63.96/27'):[],
    ipaddress.ip_network('172.21.63.128/26'):[],
    ipaddress.ip_network('172.21.63.192/26'):[],
    ipaddress.ip_network('172.21.66.0/25'):[],
    ipaddress.ip_network('172.21.67.0/24'):[],
    ipaddress.ip_network('172.21.88.0/24'):[],
    ipaddress.ip_network('172.21.196.0/24'):[],
    ipaddress.ip_network('172.21.197.0/30'):[],
    ipaddress.ip_network('172.21.197.4/30'):[],
    ipaddress.ip_network('172.21.197.8/29'):[],
    ipaddress.ip_network('172.21.208.0/20'):[]
}

# Custom subnet address management for 172.16.xx.xx/12 SJC23
subnets_172_25_xx_xx = {
    ipaddress.ip_network('172.25.192.0/24'):[],
    ipaddress.ip_network('172.25.193.0/24'):[],
    ipaddress.ip_network('172.25.194.0/24'):[],
    ipaddress.ip_network('172.25.195.0/24'):[],
    ipaddress.ip_network('172.25.196.0/24'):[],
    ipaddress.ip_network('172.25.197.0/24'):[],
    ipaddress.ip_network('172.25.198.0/24'):[],
    ipaddress.ip_network('172.25.199.0/24'):[]
}

# Custom subnet address management for 192.168.xx.xx/20 SJC23
subnets_192_168_xx_xx = {
    ipaddress.ip_network('192.168.224.0/24'):[],
    ipaddress.ip_network('192.168.225.0/24'):[],
    ipaddress.ip_network('192.168.226.0/24'):[],
    ipaddress.ip_network('192.168.227.0/24'):[],
    ipaddress.ip_network('192.168.228.0/24'):[],
    ipaddress.ip_network('192.168.229.0/24'):[],
    ipaddress.ip_network('192.168.230.0/24'):[],
    ipaddress.ip_network('192.168.231.0/24'):[],
    ipaddress.ip_network('192.168.232.0/24'):[],
    ipaddress.ip_network('192.168.233.0/24'):[],
    ipaddress.ip_network('192.168.234.0/24'):[],
    ipaddress.ip_network('192.168.235.0/24'):[],
    ipaddress.ip_network('192.168.236.0/24'):[],
    ipaddress.ip_network('192.168.237.0/24'):[],
    ipaddress.ip_network('192.168.238.0/25'):[],
    ipaddress.ip_network('192.168.238.128/28'):[],
    ipaddress.ip_network('192.168.238.144/28'):[],
    ipaddress.ip_network('192.168.238.160/27'):[],
    ipaddress.ip_network('192.168.238.192/27'):[],
    ipaddress.ip_network('192.168.238.224/28'):[],
    ipaddress.ip_network('192.168.238.240/28'):[],
    ipaddress.ip_network('192.168.239.0/24'):[]
}

subnets_doods = {
    ipaddress.ip_network('10.22.0.0/16'):subnets_10_22_xx_xx,
    ipaddress.ip_network('10.29.192.0/18'):subnets_10_29_xx_xx,
    ipaddress.ip_network('171.68.160.0/19'):subnets_171_68_xx_xx,
    ipaddress.ip_network('172.21.0.0/16'):subnets_172_21_xx_xx,
    ipaddress.ip_network('172.25.0.0/16'):subnets_172_25_xx_xx,
    ipaddress.ip_network('192.168.224.0/20'):subnets_192_168_xx_xx
}

class Vulnerability():
    '''
    A class to store vulnerabilities and their useful information
    '''
    def __init__(self,key='',summary='',lab_id=0,ip=0, priority='', description=''):
        self.key = key
        self.summary = summary
        self.lab_id = lab_id
        self.ip = ipaddress.ip_address(ip)
        self.priority = priority
        self.description = description

    def __str__(self):
        return f'Key: {self.key}\tSummary: {self.summary}\n\tLab ID: {self.lab_id}\tIP: {self.ip}  \tPriority: {self.priority}'

def open_file():
    '''
    Prompt user for vulnerability file
    '''
    while True:
        try:
            infile = input('Enter full path of vulnerability file: ')
            fh = open(infile, 'r')
        except FileNotFoundError:
            print("File not found.")
            continue
        except PermissionError:
            print("Permission denied.")
            continue
        except OSError:
            print("Invalid argument.")
            continue
        else:
            fh.close()
            break
    return infile

def parse_file(ifile):
    '''
    Input: A raw text file of all vulnerability information
    Output: A list of vunerabilities grouped by key
    '''
    # empty object list to return
    long_list = []
    # flag to check for garbage text
    vuln_starts = False

    # read the file line by line and save into list
    with open(ifile, 'r') as fh:
        text = fh.readlines()
    for line in text:
        # splitting the list into categories based on tab character
        categories = re.split(r'\t+', line.rstrip('\n'))
        # create a long_list that is divided by \n and \t
        for section in categories:
            if section[0:5] == 'CLVM-':
                # set escape character to separate vulnerabilities
                long_list.append('||')
                vuln_starts = True
            if not vuln_starts:
                continue
            long_list.append(section)
    # split list into sections per vulnerability based on escape character '||'
    vulnerabilities = [list(group) for k, group in groupby(long_list, lambda x: x == "||") if not k]

    return vulnerabilities

def create_useful_vulnerability_list(vulnerability_list):
    '''
    This function can be changed if more information about each vulnerability is needed
    Input: A list of all vulnerabilities grouped by key
    Output: List of necessary vulnerability information
    '''
    # create an empty list of vulnerabilities
    # NOTE: some entries have a port # and others do not
    vulnerabilities = []
    for vulnerability in vulnerability_list:
        # if port number is included in entry, skip the port number and add IP address at [4]
        if '.' not in vulnerability[3]:
            vuln = Vulnerability(vulnerability[0],vulnerability[1],vulnerability[2],vulnerability[4], vulnerability[7], '\n'.join(vulnerability[8:-3]))
            vulnerabilities.append(vuln)
        # no port number is included in entry, add IP address at [3]
        else:
            vuln = Vulnerability(vulnerability[0],vulnerability[1],vulnerability[2],vulnerability[3], vulnerability[6],'\n'.join(vulnerability[7:-3]))
            vulnerabilities.append(vuln)
    # a full useful list of vulnerabilities is returned
    return vulnerabilities

def populate_subnets(vulnerability_list):
    '''
    This will place each vulnerability object in their respective subnet based on IP address
    '''
    global count
    for vulnerability in vulnerability_list:
        for parent_subnet in subnets_doods:
            if vulnerability.ip in parent_subnet:
                for subnet in subnets_doods[parent_subnet]:
                    if vulnerability.ip in subnet:
                         subnets_doods[parent_subnet][subnet].append(vulnerability)
                         count += 1
                         break

def select_subnet(rank):
    '''
    Prompt user for subnet
    Input: level of subnet to be selected
    Output: valid subnet
    '''
    while True:
        valid_subnet = False
        try:
            subnet = ipaddress.ip_network(input('Please input a subnet: '))
        except KeyboardInterrupt:
            return False
        except:
            print('That is not a valid subnet format!')
            continue
        else:
            if rank == 'parent':
                if subnets_doods.get(subnet):
                    valid_subnet = True
            elif rank == 'child':
                for parent_subnet in subnets_doods:
                    if subnets_doods[parent_subnet].get(subnet):
                        valid_subnet = True
                        break
            if valid_subnet:
                break
            else:
                print('This subnet does not have vulnerabilities OR it is an invalid subnet choice.')
    return subnet

def sort_by_priority():
    for parent_subnet in subnets_doods:
        for subnet in subnets_doods[parent_subnet]:
            if subnets_doods[parent_subnet].get(subnet):
                subnets_doods[parent_subnet][subnet].sort(key=lambda x: x.priority)

def print_all():
    '''
    Print all of the vulnerabilities sorted with their appropriate subnet
    '''
    for parent_subnet in subnets_doods:
        for subnet in subnets_doods[parent_subnet]:
            if subnets_doods[parent_subnet][subnet]:
                print(subnet)
                for vuln in subnets_doods[parent_subnet][subnet]:
                    print(vuln)
                print('\n')

def print_vulnerable_subnets():
    '''
    Print all of the subnets with vulnerabilities
    '''
    print('All vulnerable subnets:')
    for parent_subnet in subnets_doods:
        for subnet in subnets_doods[parent_subnet]:
            if subnets_doods[parent_subnet].get(subnet):
                print(subnet)

def print_parent_subnets():
    '''
    Print the top layer subnets
    '''
    print('\nParent subnets:')
    for parent_subnet in subnets_doods:
        local_count = 0
        for subnet in subnets_doods[parent_subnet]:
            if subnets_doods[parent_subnet].get(subnet):
                local_count += len(subnets_doods[parent_subnet][subnet])
        print(f'{parent_subnet}     \t{local_count} vulnerabilities')

def expand_subnet(parent_subnet):
    '''
    Print the children subnets of the top layer subnet
    '''
    print(f'\nSubnets with vulnerabilities from {parent_subnet}')
    for subnet in subnets_doods[parent_subnet]:
            if subnets_doods[parent_subnet].get(subnet):
                print(f'{subnet}    \t{len(subnets_doods[parent_subnet][subnet])} vulnerabilities')

def print_vulnerable_ips_in_subnet(child_subnet):
    '''
    Print the vulnerabilities within a specific subnet
    '''
    print(f'\nVulnerabilities in {child_subnet}')
    for parent_subnet in subnets_doods:
        if subnets_doods[parent_subnet].get(child_subnet):
            for vuln in subnets_doods[parent_subnet][child_subnet]:
                print(vuln)
            print(f'\nThere are: {len(subnets_doods[parent_subnet][child_subnet])} vulnerabilities in this subnet.\n')
            break

def export_vulnerabilities_to_excel(child_subnet):
    global path
    # creating workbook
    workbook = Workbook()
    ws = workbook.active

    # adding bold text style
    bold = Font(bold=True)

    # set dimensions of columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 90
    ws.column_dimensions['F'].width = 100

    # specifying column names
    ws['A1'] = 'Key'
    ws['B1'] = 'IP Address'
    ws['C1'] = 'Lab ID'
    ws['D1'] = 'InfoSec Severity'
    ws['E1'] = 'Summary'
    ws['F1'] = 'Description'

    # apply bold font to row 1
    for cell in ws["1:1"]:
        cell.font = bold

    # populate excel sheet with vulnerabilities
    row = 2
    for parent_subnet in subnets_doods:
        if subnets_doods[parent_subnet].get(child_subnet):
            for vuln in subnets_doods[parent_subnet][child_subnet]:
                ws['A'+str(row)] = str(vuln.key)
                ws['B'+str(row)] = str(vuln.ip)
                ws['C'+str(row)] = str(vuln.lab_id)
                ws['D'+str(row)] = str(vuln.priority)
                ws['E'+str(row)] = str(vuln.summary)
                ws['F'+str(row)] = str(vuln.description)
                row += 1

    # prompt user for file name
    filename = input('Enter name of file to save: ')
    filename += '.xlsx'

    # save the workbook
    #workbook.save(path+filename)
    workbook.save(filename)

def menu():
    '''
    Dedicated menu for searching through subnets/ips
    '''
    while True:
        print_parent_subnets()
        parent_subnet = select_subnet('parent')
        if not parent_subnet:
            break
        while parent_subnet:
            expand_subnet(parent_subnet)
            # add choice to go back or select child subnet
            back = input('Select child subnet (1)\tBack to parent subnets (2): ')
            # input checking
            while back not in ('1','2'):
                back = input('Please select (1) or (2) from above: ')
            if back == '2':
                break
            else:
                child_subnet = select_subnet('child')
                print_vulnerable_ips_in_subnet(child_subnet)
                export_vulnerabilities_to_excel(child_subnet)
                # press enter to continue (back to expanded subnet)
                (input('Press Enter to continue...'))


def main():
    '''
    Main program running
    '''
    global path
    print("Welcome to the Vulnerability Organizer beta!\n")
    # check if file is valid
    infile = open_file()
    # separate file into groups of information
    vulnerability_list = parse_file(infile)
    # create useful list of vulnerabilities from groups of information
    vulnerability_list = create_useful_vulnerability_list(vulnerability_list)
    # sort the useful list by IP address
    vulnerability_list.sort(key=lambda x: x.ip)
    
    # ************Uncomment this if missing vulnerabilities/subnets****************
    #for vulnerability in vulnerability_list:
    #    print(vulnerability.ip)

    # populate global dictionaries with vulnerabilities from list
    populate_subnets(vulnerability_list)
    # sort each subnet by vulnerability priority
    sort_by_priority()
    print("\nData has been populated and organized!")
    print(f'There are: {len(vulnerability_list)} vulnerabilities in the file given.')
    print(f'There are: {count} vulnerabilities accounted for. Subnets for missing vulnerabilities have not been added yet.')
    input('Press Enter to continue...')

    # will work on path later (issue with path verification)
    #path = input('\nEnter path to save files to: ')
    menu()

if __name__ == "__main__":
    main()